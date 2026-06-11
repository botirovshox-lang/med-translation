"""
recover_fp.py — Восстановление ложно удалённых записей (false positives из tilde-фильтра).

Проблема: is_bad_tilde_artifact() в fix_issues.py был слишком агрессивным:
  - удалял gluten-free, self-acting, allergen-challenged (гифенированные прилагательные)
  - удалял adventitious, adventive (синонимы через запятую)
  - удалял antibiotic-resistant (одно гифенированное слово)

Решение: улучшенный детектор + восстановление ложно удалённых записей из *_fix_rejected.tsv.

После восстановления: регенерация всех XLSX.
"""
import sys, csv, re, os
from collections import OrderedDict
sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

FINAL = r"C:\Users\Shox\med_translation\final_output"


# ─────────────────────────────────────────────────────────────────────────────
# УЛУЧШЕННЫЙ ДЕТЕКТОР АРТЕФАКТОВ ТИЛЬДЫ
# ─────────────────────────────────────────────────────────────────────────────

_ADJ_RU_PAT = re.compile(
    r'(?:ский|кий|ный|ной|жий|хий|ший|зий|дний|зний|вый|лый|гий|бый|пый|рый|дый|тый|зый|фый|нный)$',
    re.IGNORECASE
)

# Суффиксы английских прилагательных и причастий
_ADJ_EN_ENDINGS = (
    'ous', 'al', 'ic', 'ive', 'ary', 'ory', 'ant', 'ent', 'ful', 'less',
    'ish', 'ile', 'ine', 'ble', 'ical', 'ed', 'ated', 'ized', 'ised',
    # устойчивые составные прилагательные:
    'free', 'prone', 'based', 'related', 'mediated', 'challenged',
    'induced', 'linked', 'dependent', 'associated', 'derived',
    'acting', 'forming', 'binding', 'sensing', 'nourishing', 'resistant',
    'positive', 'negative', 'specific', 'reactive', 'sensitive',
)

_STOPWORDS = frozenset({
    'a','an','the','of','in','at','to','by','for','and','or','as',
    'is','are','with','its','their'
})

# Суффиксы-маркеры множественного числа/вариантов в словаре
_PLURAL_MARKER = re.compile(
    r',\s*(?:pl|pi|sing|abbr|var|obs|syn)\.\s*\S+', re.I
)


def is_single_adjective(ru: str) -> bool:
    words = ru.strip().split()
    return len(words) == 1 and bool(_ADJ_RU_PAT.search(words[0].rstrip('.,;:')))


def is_bad_tilde_artifact_v2(ru: str, en: str) -> bool:
    """
    Улучшенный детектор: возвращает True ТОЛЬКО для настоящих артефактов тильды
    (однословное RU-прилагательное + EN составная именная фраза).

    Сохраняет:
      - запятые-синонимы:  adventitious, adventive       → False (keep)
      - гифенированные:    antibiotic-resistant           → False (keep)
      - все токены с гифеном: immune-mediated            → False (keep)
      - adj-окончания:     affinity purified              → False (keep, 'ed')
      - аббревиатуры:      chronic active (CAH)           → False (keep, 'CAH'≤3)

    Удаляет:
      - aberrant goiter    → True  (goiter = noun, нет adj-суффикса)
      - albuminous periostitis → True  (periostitis = noun)
      - auriculotemporal nerve syndrome → True (syndrome = noun)
      - antibody-induced capping → True (capping после гифенного токена = noun)
    """
    if not is_single_adjective(ru):
        return False

    # ── 1. Множественные синонимы через запятую (не маркер мн.ч.) ──────────
    if ',' in en:
        cleaned = _PLURAL_MARKER.sub('', en)
        if ',' in cleaned:
            return False   # реальные синонимы → оставляем

    # ── 2. Убираем скобочные аббревиатуры: "chronic active (CAH)" → "chronic active"
    en_core = re.sub(r'\s*\([A-Z]{1,6}\)\s*', ' ', en).strip()

    # ── 3. Токены (пробел-разделение)
    tokens = [t for t in en_core.split() if re.search(r'[a-zA-Z]', t)]

    # Убираем все-заглавные аббревиатуры
    non_abbrev = [t for t in tokens if not (t.isupper() and len(t) <= 5)]
    if not non_abbrev:
        return False

    # ── 4. Все токены гифенированные → составное прилагательное, оставляем ─
    if all('-' in t for t in non_abbrev):
        return False

    # ── 5. Собираем все буквенные слова (кроме стоп-слов) ──────────────────
    words = [w for w in re.findall(r'[a-zA-Z]{2,}', en_core)
             if w.lower() not in _STOPWORDS]
    if len(words) <= 1:
        return False

    # ── 6. Проверяем ПОСЛЕДНЕЕ значимое слово ──────────────────────────────
    # Пропускаем аббревиатуры (ERCP, CAH…) в конце
    sig_words = [w for w in words if not (w.isupper() and len(w) <= 5)]
    if not sig_words:
        return False

    last = sig_words[-1].lower()

    # Очень короткое последнее слово → скорее всего не существительное
    if len(last) <= 3:
        return False

    # Последнее слово с прилагательным суффиксом → не артефакт
    if any(last.endswith(e) for e in _ADJ_EN_ENDINGS):
        return False

    # Иначе → настоящий артефакт тильды (существительная фраза)
    return True


# ─────────────────────────────────────────────────────────────────────────────
# ВОССТАНОВЛЕНИЕ ЛОЖНЫХ ПОЗИТИВОВ
# ─────────────────────────────────────────────────────────────────────────────

def recover_file(main_tsv, rejected_tsv, ru_col, en_col, label):
    """
    1. Читаем основной TSV и rejected TSV.
    2. Для каждого tilde_artifact в rejected: применяем новый детектор.
       Если entry теперь NOT bad → восстанавливаем.
    3. Обновляем EN в основном файле (если RU уже есть → дополняем, нет → новая строка).
    4. Сохраняем обновлённый TSV.
    """
    # Читаем основной файл
    if not os.path.exists(main_tsv):
        print(f"  ⚠  Не найден: {main_tsv}")
        return None

    with open(main_tsv, encoding='utf-8-sig') as f:
        main_rows = list(csv.DictReader(f, delimiter='\t'))
    fields = list(main_rows[0].keys()) if main_rows else []
    total_before = len(main_rows)

    # Строим индекс по нормализованному RU: ru_lower → индекс в main_rows
    ru_index = {}
    for i, r in enumerate(main_rows):
        key = r.get(ru_col, '').lower().strip()
        if key not in ru_index:
            ru_index[key] = i

    # Читаем rejected
    if not os.path.exists(rejected_tsv):
        print(f"  ⚠  Rejected-файл не найден: {rejected_tsv}")
        return {'rows': main_rows, 'fields': fields,
                'recovered': 0, 'total_in': total_before}

    with open(rejected_tsv, encoding='utf-8-sig') as f:
        rejected_rows = list(csv.DictReader(f, delimiter='\t'))

    # Фильтруем: только tilde_artifact
    tilde_rej = [r for r in rejected_rows
                 if 'tilde_artifact' in r.get('_reject_reason', '')]

    # Группируем по RU (несколько EN для одного RU)
    by_ru = OrderedDict()
    for r in tilde_rej:
        key = r.get(ru_col, '').lower().strip()
        if key not in by_ru:
            by_ru[key] = []
        by_ru[key].append(r)

    recovered_count = 0
    still_bad_count = 0
    recovered_examples = []

    for ru_key, group in by_ru.items():
        # Проверяем, является ли ХОТЬ ОДИН из EN ложным позитивом
        fp_rows = []
        for r in group:
            ru = r.get(ru_col, '').strip()
            en = r.get(en_col, '').strip()
            if not is_bad_tilde_artifact_v2(ru, en):
                fp_rows.append(r)
            else:
                still_bad_count += 1

        if not fp_rows:
            continue

        # Собираем уникальные EN из ложных позитивов
        new_en_parts = []
        seen_en_norm = set()
        for r in fp_rows:
            en = r.get(en_col, '').strip()
            en_norm = re.sub(r'[^a-z0-9]', '', en.lower())
            if en_norm and en_norm not in seen_en_norm:
                seen_en_norm.add(en_norm)
                new_en_parts.append(en)

        if not new_en_parts:
            continue

        recovered_count += len(new_en_parts)
        recovered_examples.extend(fp_rows[:2])

        if ru_key in ru_index:
            # RU уже есть — дополняем EN
            idx = ru_index[ru_key]
            existing_en = main_rows[idx].get(en_col, '')
            # Проверяем дубли
            existing_norm = set(
                re.sub(r'[^a-z0-9]', '', part.lower().strip())
                for part in existing_en.split(';') if part.strip()
            )
            unique_new = [e for e in new_en_parts
                         if re.sub(r'[^a-z0-9]', '', e.lower()) not in existing_norm]
            if unique_new:
                all_en = [existing_en] + unique_new if existing_en else unique_new
                main_rows[idx][en_col] = '; '.join(all_en)
        else:
            # RU не существует — добавляем новую строку
            base = dict(fp_rows[0])
            # Убираем служебное поле
            base.pop('_reject_reason', None)
            base[en_col] = '; '.join(new_en_parts)
            # Убираем лишние поля
            clean_base = {k: base.get(k, '') for k in fields}
            main_rows.append(clean_base)
            # Добавляем в индекс
            ru_index[ru_key] = len(main_rows) - 1

    # Сортируем по RU
    main_rows.sort(key=lambda r: r.get(ru_col, '').lower())

    # Сохраняем
    with open(main_tsv, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter='\t', extrasaction='ignore')
        w.writeheader()
        w.writerows(main_rows)

    print(f"\n{'─'*62}")
    print(f"  {label}")
    print(f"{'─'*62}")
    print(f"  Строк до восстановления:   {total_before:>8,}")
    print(f"  Tilde-rejected записей:    {len(tilde_rej):>8,}")
    print(f"  Истинные артефакты (ок):   {still_bad_count:>8,}")
    print(f"  Ложные позитивы (возврат): {recovered_count:>8,}")
    print(f"  Строк после восстановления:{len(main_rows):>8,}")

    if recovered_examples:
        print(f"\n  Примеры восстановленных (первые 8):")
        for r in recovered_examples[:8]:
            print(f"    RU: {r.get(ru_col,'')[:45]}")
            print(f"    EN: {r.get(en_col,'')[:55]}")

    return {'rows': main_rows, 'fields': fields,
            'recovered': recovered_count, 'total_in': len(main_rows)}


# ─────────────────────────────────────────────────────────────────────────────
# XLSX-ГЕНЕРАЦИЯ (с обработкой PermissionError)
# ─────────────────────────────────────────────────────────────────────────────

def save_xlsx_safe(rows, col_headers, xlsx_path, sheet_name, hdr_color):
    """Сохраняет XLSX. При PermissionError пробует альтернативное имя."""
    if not HAS_XLSX or not rows:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    thin = Side(style='thin', color='CCCCCC')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Заголовок
    for ci, col in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font  = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        cell.fill  = PatternFill(fill_type='solid', fgColor=hdr_color)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bdr
    ws.row_dimensions[1].height = 20

    # Данные
    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(col_headers, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col, ''))
            cell.alignment = Alignment(vertical='top', wrap_text=False)
            cell.border = bdr
            if ri % 2 == 0:
                cell.fill = PatternFill(fill_type='solid', fgColor='F2F7FF')

    # Ширина колонок
    col_widths = {c: len(c) + 2 for c in col_headers}
    for row in rows[:500]:
        for col in col_headers:
            col_widths[col] = min(70, max(col_widths[col], len(str(row.get(col, ''))) + 2))
    for ci, col in enumerate(col_headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[col]

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # Попытка сохранения
    for attempt, path in enumerate([xlsx_path, xlsx_path.replace('.xlsx', '_v2.xlsx')]):
        try:
            wb.save(path)
            name = os.path.basename(path)
            if attempt > 0:
                print(f"  ✓ {name} → {len(rows):,} строк  (файл был занят, сохранено под v2)")
            else:
                print(f"  ✓ {name} → {len(rows):,} строк")
            return
        except PermissionError:
            if attempt == 0:
                print(f"  ⚠  {os.path.basename(path)} занят (закрой Excel), пробуем _v2...")
            else:
                print(f"  ✗  Не удалось сохранить {os.path.basename(path)} — закрой Excel")


# ─────────────────────────────────────────────────────────────────────────────
# ОСНОВНАЯ ЛОГИКА
# ─────────────────────────────────────────────────────────────────────────────

print("═" * 65)
print("RECOVER FALSE POSITIVES — восстановление ложно удалённых записей")
print("═" * 65)

results = {}

results['ref'] = recover_file(
    main_tsv     = FINAL + r'\reference_glossary_FINAL.tsv',
    rejected_tsv = FINAL + r'\reference_glossary_FINAL_fix_rejected.tsv',
    ru_col       = 'Russian',
    en_col       = 'English',
    label        = 'reference_glossary_FINAL',
)

results['app'] = recover_file(
    main_tsv     = FINAL + r'\approved_glossary_FINAL.tsv',
    rejected_tsv = FINAL + r'\approved_glossary_FINAL_fix_rejected.tsv',
    ru_col       = 'Russian',
    en_col       = 'English',
    label        = 'approved_glossary_FINAL',
)

results['lkp'] = recover_file(
    main_tsv     = FINAL + r'\akzhigitov_lookup.tsv',
    rejected_tsv = FINAL + r'\akzhigitov_lookup_fix_rejected.tsv',
    ru_col       = 'Russian',
    en_col       = 'English',
    label        = 'akzhigitov_lookup',
)


# ─────────────────────────────────────────────────────────────────────────────
# ГЕНЕРАЦИЯ XLSX
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "═" * 65)
print("ГЕНЕРАЦИЯ XLSX")
print("═" * 65)

# Reference — зелёная тема
if HAS_XLSX and results.get('ref'):
    r = results['ref']
    ref_cols = [f for f in r['fields'] if f != '_reject_reason']
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reference'
    thin = Side(style='thin', color='CCCCCC')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(ref_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font  = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        cell.fill  = PatternFill(fill_type='solid', fgColor='145A32')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 20
    for ri, row in enumerate(r['rows'], 2):
        for ci, col in enumerate(ref_cols, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col, ''))
            cell.alignment = Alignment(vertical='top')
            if ri % 2 == 0:
                cell.fill = PatternFill(fill_type='solid', fgColor='EAFAF1')
    col_w = {'Russian': 44, 'English': 54, 'Category': 16,
             'Confidence': 20, 'Sources': 22, 'Issues': 18}
    for ci, col in enumerate(ref_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_w.get(col, 20)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    xlsx_path = FINAL + r'\reference_glossary_FINAL.xlsx'
    for attempt, path in enumerate([xlsx_path, xlsx_path.replace('.xlsx', '_v2.xlsx')]):
        try:
            wb.save(path)
            name = os.path.basename(path)
            suffix = ' (файл был занят, сохранено _v2)' if attempt > 0 else ''
            print(f"  ✓ {name} → {r['total_in']:,} строк{suffix}")
            break
        except PermissionError:
            if attempt == 0:
                print(f"  ⚠  reference_glossary_FINAL.xlsx занят, пробуем _v2...")

# Approved
if results.get('app'):
    save_xlsx_safe(
        rows        = results['app']['rows'],
        col_headers = [f for f in results['app']['fields'] if f != '_reject_reason'],
        xlsx_path   = FINAL + r'\approved_glossary_FINAL.xlsx',
        sheet_name  = 'Approved',
        hdr_color   = '1A5276',
    )

# Lookup
if results.get('lkp'):
    save_xlsx_safe(
        rows        = results['lkp']['rows'],
        col_headers = [f for f in results['lkp']['fields'] if f != '_reject_reason'],
        xlsx_path   = FINAL + r'\akzhigitov_lookup.xlsx',
        sheet_name  = 'Lookup',
        hdr_color   = '784212',
    )


# ─────────────────────────────────────────────────────────────────────────────
# ИТОГ
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "═" * 65)
print("ИТОГ RECOVER_FP")
print("═" * 65)
for key, label in [('ref', 'reference'), ('app', 'approved'), ('lkp', 'lookup')]:
    r = results.get(key)
    if r:
        print(f"  {label:<20} {r['total_in']:>7,} строк  (+{r['recovered']:,} восстановлено)")

total = sum(results[k]['total_in'] for k in results if results.get(k))
print(f"\n  Итого записей: {total:,}")
print(f"  Файлы обновлены: {FINAL}")

print("\n  Детальная проверка улучшенного детектора:")
test_cases = [
    # (ru, en, should_be_deleted)
    ('хронический',        'chronic pyelonephritis',       True),
    ('хронический',        'chronic active (CAH)',          False),  # active=adj -ive
    ('аберрантный',        'aberrant goiter',               True),
    ('аглютеновый',        'gluten-free',                   False),
    ('автоматический',     'self-acting',                   False),
    ('антибиотикорезистентный', 'antibiotic-resistant',     False),
    ('аллергизированный',  'allergen-challenged',           False),
    ('аллергочувствительный', 'allergy-prone',              False),
    ('бессимптомный',      'symptom-free',                  False),
    ('адвентициальный',    'adventitious, adventive',       False),
    ('актиноморфный',      'actinomorphic, actinomorphous', False),
    ('альбуминозный',      'albuminous periostitis',        True),
    ('аурикулотемпоральный','auriculotemporal nerve syndrome', True),
    ('аутоиммунный',       'immune-mediated keratoconjunctivitis', True),
    ('аутоиммунный',       'immune-mediated',               False),
    ('аффинный',           'affinity purified',             False),
]
ok = 0
fail = 0
for ru, en, expected_del in test_cases:
    got = is_bad_tilde_artifact_v2(ru, en)
    status = '✓' if got == expected_del else '✗'
    if got == expected_del:
        ok += 1
    else:
        fail += 1
    print(f"    {status} is_bad({ru!r}, {en!r}) = {got}  (ожид.{expected_del})")
print(f"\n  Тестов: {len(test_cases)}  ✓ {ok}  ✗ {fail}")
