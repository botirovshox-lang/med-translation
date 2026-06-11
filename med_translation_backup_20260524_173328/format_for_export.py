"""
Финальное форматирование для экспорта в CAT-инструменты:
  1. Убираем записи с «см.» из reference_glossary_FINAL.tsv
  2. Чистим TM: убираем английские заголовки из начала RU-сегментов
  3. Создаём tm_reference_FINAL.tmx (TMX 1.4) для Trados/OmegaT/MemoQ
  4. Создаём XLSX-версии approved и reference глоссариев
  5. Создаём 2_5_glossary.xlsx и covid_glossary.xlsx из parsed пар
"""
import sys, csv, re, os
from datetime import datetime, timezone
sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    print("⚠  openpyxl не установлен — XLSX не будут созданы.")
    print("   Установите: pip install openpyxl")
    HAS_OPENPYXL = False

PARSED = r"C:\Users\Shox\med_translation\словари\parsed"
FINAL  = r"C:\Users\Shox\med_translation\final_output"

# ─────────────────────────────────────────────────────────────────────────────
# Утилиты
# ─────────────────────────────────────────────────────────────────────────────

def has_cyrillic(s):
    return bool(re.search(r"[А-ЯЁа-яё]", s))

def cyrillic_ratio(s):
    if not s:
        return 0.0
    letters = re.findall(r"[A-Za-zА-ЯЁа-яё]", s)
    if not letters:
        return 0.0
    cyr = sum(1 for c in letters if re.match(r"[А-ЯЁа-яё]", c))
    return cyr / len(letters)

def is_see_ref(ru):
    """Запись является чистой отсылкой «см.» без реального перевода."""
    ru = ru.strip()
    if not ru:
        return True
    # «см.» в начале → нет перевода
    if re.match(r"^см\b", ru, re.IGNORECASE):
        return True
    # Запись ТОЛЬКО из «см.» + что-то, без кириллицы до «см.»
    m = re.match(r"^(.*?)\s*см\.\s*(.*)$", ru, re.IGNORECASE)
    if m:
        before = m.group(1).strip()
        # Если до «см.» нет содержательного кириллического слова
        cyr_before = re.findall(r"[А-ЯЁа-яё]{2,}", before)
        if not cyr_before:
            return True
    return False


# ─────────────────────────────────────────────────────────────────────────────
# 1. Чистим reference от «см.» записей
# ─────────────────────────────────────────────────────────────────────────────
print("═" * 65)
print("1. УДАЛЕНИЕ ЗАПИСЕЙ «СМ.» ИЗ reference_glossary_FINAL.tsv")
print("═" * 65)

ref_path = FINAL + r"\reference_glossary_FINAL.tsv"
with open(ref_path, encoding="utf-8-sig") as f:
    ref_all = list(csv.DictReader(f, delimiter="\t"))

ref_fields = ["Russian", "English", "Category", "Confidence", "Sources", "Issues"]

ref_clean = []
ref_removed = []
for r in ref_all:
    ru = r.get("Russian", "").strip()
    if is_see_ref(ru):
        ref_removed.append(r)
    else:
        ref_clean.append(r)

print(f"  Было:    {len(ref_all):,}")
print(f"  Удалено: {len(ref_removed):,}")
for r in ref_removed:
    print(f"    EN: {r['English'][:50]}  →  RU: {r['Russian'][:60]}")
print(f"  Стало:   {len(ref_clean):,}")

with open(ref_path, "w", encoding="utf-8-sig", newline="") as f:
    w = csv.DictWriter(f, fieldnames=ref_fields, delimiter="\t", extrasaction="ignore")
    w.writeheader()
    w.writerows(ref_clean)
print(f"  ✓ reference_glossary_FINAL.tsv обновлён\n")


# ─────────────────────────────────────────────────────────────────────────────
# 2. Чистим TM-сегменты
# ─────────────────────────────────────────────────────────────────────────────
print("═" * 65)
print("2. ОЧИСТКА TM-СЕГМЕНТОВ")
print("═" * 65)

tm_path = FINAL + r"\tm_reference_FINAL.tsv"

try:
    with open(tm_path, encoding="utf-8-sig") as f:
        tm_all = list(csv.DictReader(f, delimiter="\t"))
    print(f"  TM загружен: {len(tm_all):,} сегментов")
    tm_found = True
except FileNotFoundError:
    print(f"  ⚠ Файл {tm_path} не найден — TM-шаг пропущен")
    tm_all = []
    tm_found = False

# Паттерн: одно или два английских слова в начале RU-сегмента (артефакт MedlinePlus)
# Пример: "мочи Test Суточный анализ" → убрать "Test "
# Пример: "Test Суточный" → убрать "Test "
# Пример: "Pregnancy Беременность" → убрать "Pregnancy "
EN_AT_START = re.compile(
    r"^(?:[A-Z][a-z]*(?:\s+[A-Z][a-z]*)?\s+)+"  # Одно+ капитальных EN-слов
)
EN_EMBEDDED = re.compile(
    r"\s+[A-Z][a-z]{2,}(?:\s+[A-Z][a-z]{2,})*\s+"  # EN слова внутри строки
)

def clean_tm_ru(s):
    """Убираем английские заголовки из начала и середины RU-сегмента."""
    s = s.strip()
    # Убираем EN-слова в самом начале (артефакты MedlinePlus)
    m = EN_AT_START.match(s)
    if m:
        # Проверяем: после удаления остаётся кириллица?
        tail = s[m.end():]
        if has_cyrillic(tail):
            s = tail.strip()
    # Нормализуем пробелы
    s = re.sub(r"\s+", " ", s).strip()
    return s

if tm_found and tm_all:
    # Проверяем поля
    sample_keys = list(tm_all[0].keys())
    print(f"  Поля TM: {sample_keys}")

    # Угадываем имена полей
    ru_field = next((k for k in sample_keys if "uss" in k or k == "RU" or k.lower() == "russian"), None)
    en_field = next((k for k in sample_keys if "ngl" in k or k == "EN" or k.lower() == "english"), None)

    if not ru_field or not en_field:
        # Пробуем первые два поля
        if len(sample_keys) >= 2:
            ru_field, en_field = sample_keys[0], sample_keys[1]

    print(f"  RU-поле: {ru_field!r}  EN-поле: {en_field!r}")

    tm_clean = []
    tm_skipped = 0
    for r in tm_all:
        ru_orig = r.get(ru_field, "").strip()
        en_orig = r.get(en_field, "").strip()

        if not ru_orig or not en_orig:
            tm_skipped += 1
            continue

        ru_fixed = clean_tm_ru(ru_orig)

        # Фильтрация по кириллическому соотношению (строго)
        if cyrillic_ratio(ru_fixed) < 0.75:
            tm_skipped += 1
            continue

        # Минимальная длина
        if len(ru_fixed) < 15 or len(en_orig) < 10:
            tm_skipped += 1
            continue

        entry = dict(r)
        entry[ru_field] = ru_fixed
        tm_clean.append(entry)

    print(f"  Было:    {len(tm_all):,}")
    print(f"  Удалено: {tm_skipped:,}")
    print(f"  Стало:   {len(tm_clean):,}")

    # Сохраняем чистый TSV
    with open(tm_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=sample_keys, delimiter="\t", extrasaction="ignore")
        w.writeheader()
        w.writerows(tm_clean)
    print(f"  ✓ tm_reference_FINAL.tsv обновлён")

    # Показываем первые 3 чистых сегмента
    print("\n  === Первые 3 чистых RU-сегмента ===")
    for seg in tm_clean[:3]:
        print(f"  RU: {seg[ru_field][:90]}")
        print(f"  EN: {seg[en_field][:90]}")
        print()


# ─────────────────────────────────────────────────────────────────────────────
# 3. Создаём TMX 1.4
# ─────────────────────────────────────────────────────────────────────────────
print("═" * 65)
print("3. СОЗДАНИЕ TMX 1.4")
print("═" * 65)

def xml_escape(s):
    return (s.replace("&", "&amp;")
             .replace("<", "&lt;")
             .replace(">", "&gt;")
             .replace('"', "&quot;")
             .replace("'", "&apos;"))

tmx_path = FINAL + r"\tm_reference_FINAL.tmx"

if tm_found and tm_clean:
    now_utc = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")

    lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<!DOCTYPE tmx SYSTEM "tmx14.dtd">',
        '<tmx version="1.4">',
        f'  <header creationtool="med_translation_pipeline"',
        f'          creationtoolversion="1.0"',
        f'          datatype="PlainText"',
        f'          segtype="sentence"',
        f'          adminlang="ru"',
        f'          srclang="ru"',
        f'          o-tmf="med_translation"',
        f'          creationdate="{now_utc}"',
        f'  />',
        '  <body>',
    ]

    for seg in tm_clean:
        ru_val = xml_escape(seg.get(ru_field, "").strip())
        en_val = xml_escape(seg.get(en_field, "").strip())
        if not ru_val or not en_val:
            continue
        lines += [
            '    <tu>',
            f'      <tuv xml:lang="ru"><seg>{ru_val}</seg></tuv>',
            f'      <tuv xml:lang="en"><seg>{en_val}</seg></tuv>',
            '    </tu>',
        ]

    lines += ['  </body>', '</tmx>']

    with open(tmx_path, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines) + "\n")

    print(f"  ✓ tm_reference_FINAL.tmx → {len(tm_clean):,} сегментов")
    print(f"  Файл: {tmx_path}\n")
else:
    print("  ⚠ TM пустой или не найден — TMX не создан\n")


# ─────────────────────────────────────────────────────────────────────────────
# XLSX-утилиты
# ─────────────────────────────────────────────────────────────────────────────

def make_xlsx(rows, col_headers, xlsx_path, sheet_name="Glossary",
              header_color="1F4E79", freeze=True):
    """Создаёт красиво оформленный XLSX-файл."""
    if not HAS_OPENPYXL:
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Стили заголовка
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    hdr_fill  = PatternFill(fill_type="solid", fgColor=header_color)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Строка заголовков
    for ci, col_name in enumerate(col_headers, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = cell_border

    ws.row_dimensions[1].height = 20

    # Данные
    for ri, row in enumerate(rows, start=2):
        for ci, col_name in enumerate(col_headers, start=1):
            val = row.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = cell_border
            # Чередование строк
            if ri % 2 == 0:
                cell.fill = PatternFill(fill_type="solid", fgColor="F2F7FF")

    # Авто-ширина колонок
    col_widths = {col: len(col) + 2 for col in col_headers}
    for row in rows[:500]:  # семплируем первые 500 строк
        for col in col_headers:
            val = str(row.get(col, ""))
            col_widths[col] = min(60, max(col_widths[col], len(val) + 2))

    for ci, col_name in enumerate(col_headers, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[col_name]

    # Заморозка строки заголовков
    if freeze:
        ws.freeze_panes = "A2"

    # Автофильтр
    ws.auto_filter.ref = ws.dimensions

    wb.save(xlsx_path)
    return True


# ─────────────────────────────────────────────────────────────────────────────
# 4. approved_glossary_FINAL.xlsx
# ─────────────────────────────────────────────────────────────────────────────
print("═" * 65)
print("4. XLSX: approved_glossary_FINAL")
print("═" * 65)

approved_path = FINAL + r"\approved_glossary_FINAL.tsv"
try:
    with open(approved_path, encoding="utf-8-sig") as f:
        approved_rows = list(csv.DictReader(f, delimiter="\t"))

    approved_cols = ["Russian", "English", "Category", "Confidence", "Sources"]
    xlsx_approved = FINAL + r"\approved_glossary_FINAL.xlsx"

    if HAS_OPENPYXL:
        ok = make_xlsx(
            rows=approved_rows,
            col_headers=approved_cols,
            xlsx_path=xlsx_approved,
            sheet_name="Approved",
            header_color="1A5276",
        )
        if ok:
            print(f"  ✓ approved_glossary_FINAL.xlsx → {len(approved_rows):,} строк")
        else:
            print("  ✗ Ошибка создания XLSX")
    else:
        print(f"  ⚠ openpyxl не установлен — пропущено ({len(approved_rows):,} строк)")

except FileNotFoundError:
    print(f"  ⚠ Файл {approved_path} не найден")
print()


# ─────────────────────────────────────────────────────────────────────────────
# 5. reference_glossary_FINAL.xlsx
#    (большой — разбиваем на вкладки по первой букве)
# ─────────────────────────────────────────────────────────────────────────────
print("═" * 65)
print("5. XLSX: reference_glossary_FINAL")
print("═" * 65)

ref_cols = ["Russian", "English", "Category", "Confidence", "Sources", "Issues"]
xlsx_ref = FINAL + r"\reference_glossary_FINAL.xlsx"

if HAS_OPENPYXL and ref_clean:
    # Excel поддерживает 1 048 576 строк, но с 80K строк файл будет медленным.
    # Создаём одну вкладку (сортировка по RU уже есть).
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reference"

    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    hdr_fill  = PatternFill(fill_type="solid", fgColor="145A32")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col_name in enumerate(ref_cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = cell_border
    ws.row_dimensions[1].height = 20

    print(f"  Запись {len(ref_clean):,} строк в XLSX...", end=" ", flush=True)

    for ri, row in enumerate(ref_clean, start=2):
        for ci, col_name in enumerate(ref_cols, start=1):
            val = row.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="top")
            if ri % 2 == 0:
                cell.fill = PatternFill(fill_type="solid", fgColor="EAFAF1")

    # Ширина колонок
    col_widths = {"Russian": 42, "English": 42, "Category": 16,
                  "Confidence": 20, "Sources": 22, "Issues": 18}
    for ci, col_name in enumerate(ref_cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col_name, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(xlsx_ref)
    print(f"готово")
    print(f"  ✓ reference_glossary_FINAL.xlsx → {len(ref_clean):,} строк")
elif not HAS_OPENPYXL:
    print("  ⚠ openpyxl не установлен — пропущено")
else:
    print("  ⚠ ref_clean пустой")
print()


# ─────────────────────────────────────────────────────────────────────────────
# 6. 2_5_glossary.xlsx (из parsed пар Акжигитова)
# ─────────────────────────────────────────────────────────────────────────────
print("═" * 65)
print("6. XLSX: 2_5_glossary (Акжигитов)")
print("═" * 65)

akzh_tsv = PARSED + r"\2_5_pairs.tsv"
xlsx_akzh = FINAL + r"\2_5_glossary.xlsx"

try:
    with open(akzh_tsv, encoding="utf-8-sig") as f:
        akzh_rows = list(csv.DictReader(f, delimiter="\t"))

    akzh_cols = list(akzh_rows[0].keys()) if akzh_rows else ["English", "Russian"]

    if HAS_OPENPYXL:
        ok = make_xlsx(
            rows=akzh_rows,
            col_headers=akzh_cols,
            xlsx_path=xlsx_akzh,
            sheet_name="Akzhigitov",
            header_color="4A235A",
        )
        if ok:
            print(f"  ✓ 2_5_glossary.xlsx → {len(akzh_rows):,} строк")
    else:
        print(f"  ⚠ openpyxl не установлен ({len(akzh_rows):,} строк)")
except FileNotFoundError:
    print(f"  ⚠ Файл {akzh_tsv} не найден")
print()


# ─────────────────────────────────────────────────────────────────────────────
# 7. covid_glossary.xlsx
# ─────────────────────────────────────────────────────────────────────────────
print("═" * 65)
print("7. XLSX: covid_glossary")
print("═" * 65)

covid_tsv = PARSED + r"\covid_pairs.tsv"
xlsx_covid = FINAL + r"\covid_glossary.xlsx"

try:
    with open(covid_tsv, encoding="utf-8-sig") as f:
        covid_rows = list(csv.DictReader(f, delimiter="\t"))

    covid_cols = list(covid_rows[0].keys()) if covid_rows else ["English", "Russian"]

    if HAS_OPENPYXL:
        ok = make_xlsx(
            rows=covid_rows,
            col_headers=covid_cols,
            xlsx_path=xlsx_covid,
            sheet_name="COVID",
            header_color="922B21",
        )
        if ok:
            print(f"  ✓ covid_glossary.xlsx → {len(covid_rows):,} строк")
    else:
        print(f"  ⚠ openpyxl не установлен ({len(covid_rows):,} строк)")
except FileNotFoundError:
    print(f"  ⚠ Файл {covid_tsv} не найден")
print()


# ─────────────────────────────────────────────────────────────────────────────
# 8. akzhigitov_lookup.xlsx (большой словарь для программного поиска)
# ─────────────────────────────────────────────────────────────────────────────
print("═" * 65)
print("8. XLSX: akzhigitov_lookup")
print("═" * 65)

lookup_tsv  = FINAL + r"\akzhigitov_lookup.tsv"
lookup_xlsx = FINAL + r"\akzhigitov_lookup.xlsx"

try:
    with open(lookup_tsv, encoding="utf-8-sig") as f:
        lookup_rows = list(csv.DictReader(f, delimiter="\t"))

    lookup_cols = list(lookup_rows[0].keys()) if lookup_rows else ["Russian", "English", "Confidence", "Sources"]

    if HAS_OPENPYXL:
        ok = make_xlsx(
            rows=lookup_rows,
            col_headers=lookup_cols,
            xlsx_path=lookup_xlsx,
            sheet_name="Lookup",
            header_color="784212",
        )
        if ok:
            print(f"  ✓ akzhigitov_lookup.xlsx → {len(lookup_rows):,} строк")
    else:
        print(f"  ⚠ openpyxl не установлен ({len(lookup_rows):,} строк)")
except FileNotFoundError:
    print(f"  ⚠ Файл {lookup_tsv} не найден")
print()


# ─────────────────────────────────────────────────────────────────────────────
# ИТОГ
# ─────────────────────────────────────────────────────────────────────────────
print("═" * 65)
print("ИТОГ")
print("═" * 65)
print(f"""
  TSV (обновлены):
    reference_glossary_FINAL.tsv  — {len(ref_clean):,} записей (убраны «см.»)
    tm_reference_FINAL.tsv        — {len(tm_clean) if tm_found else 'N/A'} сегментов (очищены EN-заголовки)

  TMX (CAT Translation Memory):
    tm_reference_FINAL.tmx        — {len(tm_clean) if (tm_found and tm_clean) else 'N/A'} TU-юнитов

  XLSX (глоссарии для CAT / просмотра):
    approved_glossary_FINAL.xlsx  — {len(approved_rows) if 'approved_rows' in dir() else 'N/A'} строк
    reference_glossary_FINAL.xlsx — {len(ref_clean):,} строк
    2_5_glossary.xlsx             — {len(akzh_rows) if 'akzh_rows' in dir() else 'N/A'} строк  (Акжигитов полный)
    covid_glossary.xlsx           — {len(covid_rows) if 'covid_rows' in dir() else 'N/A'} строк
    akzhigitov_lookup.xlsx        — {len(lookup_rows) if 'lookup_rows' in dir() else 'N/A'} строк  (длинные пары)

  Папка: {FINAL}
""")
