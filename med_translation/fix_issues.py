"""
fix_issues.py — Финальная правка глоссарных файлов.

Проблемы (из ревью):
  1. [квадратные скобки] в EN-поле = альтернативные термины из словаря → убрать
  2. Незакрытые ( скобки в EN или RU → добавить )
  3. Однословный RU-прилагательный + многословный EN = артефакт тильды → удалить
     Пример: хронический → chronic pyelonephritis (должно быть: хронический → chronic)
  4. Дубликаты: несколько EN для одного RU → консолидировать через "; "
  5. Известные OCR-опечатки в EN → исправить
"""
import sys, csv, re, os
from collections import defaultdict, OrderedDict
sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

FINAL = r"C:\Users\Shox\med_translation\final_output"

# ─────────────────────────────────────────────────────────────────────────────
# 1. ПРАВКИ EN-ПОЛЯ
# ─────────────────────────────────────────────────────────────────────────────

# Известные OCR-опечатки в английских терминах
OCR_FIXES_EN = [
    (re.compile(r'\bastma\b',         re.I), 'asthma'),
    (re.compile(r'\bneprectomy\b',    re.I), 'nephrectomy'),
    (re.compile(r'\btybe\b',          re.I), 'type'),
    (re.compile(r'\binfluenzae\b',    re.I), 'influenzae'),   # already correct, sanity
    (re.compile(r'\bdiabetis\b',      re.I), 'diabetes'),
    (re.compile(r'\bpnemonia\b',      re.I), 'pneumonia'),
    (re.compile(r'\banaesthesia\b',   re.I), 'anesthesia'),
    (re.compile(r'\banaesthetic\b',   re.I), 'anesthetic'),
]

def apply_ocr_fixes(en):
    for pattern, replacement in OCR_FIXES_EN:
        en = pattern.sub(replacement, en)
    return en

def fix_en(en):
    """
    1. Убираем  word [synonym1, synonym2] → убираем и слово и скобки
       (паттерн: основной термин был добавлен с альтернативами в [])
    2. Убираем отдельные [квадратные скобки] если они остались
    3. Закрываем незакрытые ( )
    4. Чистим хвосты
    Возвращает (fixed_en, description).
    """
    orig = en
    changes = []

    # --- Шаг 1: убираем "слово [синонимы]" в КОНЦЕ строки ---
    # Пример: "melting curve fluorescence [fluoroscopic, fluorometric]"
    #       → "melting curve"
    m = re.search(r'\s+\S+\s*\[[^\]]+\]\s*$', en)
    if m:
        en = en[:m.start()].strip()
        changes.append(f'removed_word_brackets:{m.group().strip()!r}')

    # --- Шаг 2: убираем оставшиеся [квадратные скобки] ---
    if '[' in en:
        new_en = re.sub(r'\s*\[[^\]]*\]\s*', ' ', en).strip()
        if new_en != en:
            changes.append(f'removed_sq_brackets')
            en = new_en

    # --- Шаг 3: закрываем незакрытые () ---
    if en.count('(') > en.count(')'):
        en = en.rstrip() + ')'
        changes.append('closed_paren')

    # --- Шаг 4: убираем одиночные закрывающие скобки в начале ---
    en = re.sub(r'^\)', '', en).strip()

    # --- Шаг 5: известные опечатки ---
    en_fixed = apply_ocr_fixes(en)
    if en_fixed != en:
        changes.append('ocr_fix')
        en = en_fixed

    # --- Шаг 6: финальная чистка хвоста ---
    en = re.sub(r'[\s,;\.]+$', '', en).strip()

    return en, ('; '.join(changes) if changes else None)


def fix_ru(ru):
    """Закрываем незакрытые () в RU."""
    if ru.count('(') > ru.count(')'):
        ru = ru.rstrip() + ')'
    ru = re.sub(r'[\s,;\.]+$', '', ru).strip()
    return ru


# ─────────────────────────────────────────────────────────────────────────────
# 2. УДАЛЕНИЕ АРТЕФАКТОВ ТИЛЬДЫ (однословный RU-прилагательный → многословный EN)
# ─────────────────────────────────────────────────────────────────────────────

# Окончания русских прилагательных (мужской род)
_ADJ_PAT = re.compile(
    r'(?:ский|кий|ный|ной|жий|хий|ший|зий|дний|зний|вый|лый|гий|бый|пый|рый|дый|тый|зый|фый|нный)$',
    re.IGNORECASE
)

def is_single_adjective(ru):
    words = ru.strip().split()
    return len(words) == 1 and bool(_ADJ_PAT.search(words[0].rstrip('.,;:')))

def count_meaningful_en_words(en):
    """Считаем значимые EN-слова (длина > 1, не артикли/предлоги)."""
    STOPWORDS = {'a','an','the','of','in','at','to','by','for','and','or','as','is','are'}
    words = [w for w in re.findall(r'[a-zA-Z]{2,}', en) if w.lower() not in STOPWORDS]
    return len(words)

def is_bad_tilde_artifact(ru, en):
    """
    Однословное RU-прилагательное + многословный EN = артефакт тильды.
    Пример: хронический → chronic pyelonephritis
             (должно быть: хронический → chronic)
    """
    return is_single_adjective(ru) and count_meaningful_en_words(en) >= 2


# ─────────────────────────────────────────────────────────────────────────────
# 3. КОНСОЛИДАЦИЯ СИНОНИМОВ
# ─────────────────────────────────────────────────────────────────────────────

_CONF_RANK = {'approved': 3, 'reference_only': 2, 'needs_human_review': 1, '': 0}

def best_confidence(values):
    return max(values, key=lambda c: _CONF_RANK.get(c, 0))

def consolidate(rows, ru_col='Russian', en_col='English'):
    """
    Группируем строки по нормализованному RU, объединяем EN через '; '.
    Другие поля: Sources = объединить, Confidence = лучшее, Category = из первой строки.
    Возвращает (consolidated_rows, dedup_count).
    """
    groups = OrderedDict()
    for r in rows:
        key = r.get(ru_col, '').lower().strip()
        if key not in groups:
            groups[key] = []
        groups[key].append(r)

    result = []
    dedup_count = 0

    for key, group in groups.items():
        if len(group) == 1:
            result.append(group[0])
            continue

        # Собираем уникальные EN (нормализованные для дедупликации)
        seen_en_norm = {}
        unique_en = []
        for r in group:
            en = r.get(en_col, '').strip()
            en_norm = re.sub(r'[^a-z0-9]', '', en.lower())
            if en_norm and en_norm not in seen_en_norm:
                seen_en_norm[en_norm] = True
                unique_en.append(en)
            else:
                dedup_count += 1

        # Берём первую строку как основу
        merged = dict(group[0])
        merged[en_col] = '; '.join(unique_en)

        # Лучший уровень доверия
        merged['Confidence'] = best_confidence(
            [r.get('Confidence', '') for r in group]
        )

        # Объединяем Sources (уникальные)
        all_sources = []
        seen_src = set()
        for r in group:
            for src in r.get('Sources', '').split(';'):
                src = src.strip()
                if src and src not in seen_src:
                    seen_src.add(src)
                    all_sources.append(src)
        merged['Sources'] = '; '.join(all_sources)

        result.append(merged)

    return result, dedup_count


# ─────────────────────────────────────────────────────────────────────────────
# 4. ОСНОВНАЯ ФУНКЦИЯ ОБРАБОТКИ ФАЙЛА
# ─────────────────────────────────────────────────────────────────────────────

def process_file(in_path, out_path, ru_col='Russian', en_col='English', label=''):
    if not os.path.exists(in_path):
        print(f"  ⚠  Файл не найден: {in_path}")
        return None

    with open(in_path, encoding='utf-8-sig') as f:
        rows = list(csv.DictReader(f, delimiter='\t'))

    if not rows:
        print(f"  ⚠  Файл пустой")
        return None

    fields = list(rows[0].keys())
    total_in = len(rows)

    stats = defaultdict(int)
    clean = []
    rejected = []

    for r in rows:
        ru = r.get(ru_col, '').strip()
        en = r.get(en_col, '').strip()

        if not ru or not en:
            stats['empty'] += 1
            rejected.append(r)
            continue

        # — Исправляем RU —
        ru_new = fix_ru(ru)
        if ru_new != ru:
            r[ru_col] = ru_new
            ru = ru_new
            stats['ru_paren_fixed'] += 1

        # — Исправляем EN —
        en_new, fix_desc = fix_en(en)
        if fix_desc:
            stats[f'en_fix:{fix_desc}'] += 1
            r[en_col] = en_new
            en = en_new

        # — Удаляем если EN опустел после правок —
        if not en or len(re.sub(r'[^\w]', '', en)) < 2:
            stats['empty_after_fix'] += 1
            r['_reject_reason'] = 'empty_en_after_fix'
            rejected.append(r)
            continue

        # — Удаляем артефакты тильды (adj + compound EN) —
        if is_bad_tilde_artifact(ru, en):
            stats['bad_tilde_artifact'] += 1
            r['_reject_reason'] = f'tilde_artifact: {ru!r} → {en!r}'
            rejected.append(r)
            continue

        clean.append(r)
        stats['ok'] += 1

    # — Консолидируем синонимы —
    clean_merged, dedup_cnt = consolidate(clean, ru_col, en_col)
    stats['consolidated'] = dedup_cnt

    # — Сохраняем —
    with open(out_path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter='\t', extrasaction='ignore')
        w.writeheader()
        w.writerows(clean_merged)

    # — Сохраняем отклонённые —
    if rejected:
        reject_path = out_path.replace('.tsv', '_fix_rejected.tsv')
        reject_fields = fields + ['_reject_reason']
        with open(reject_path, 'w', encoding='utf-8-sig', newline='') as f:
            w = csv.DictWriter(f, fieldnames=reject_fields, delimiter='\t', extrasaction='ignore')
            w.writeheader()
            w.writerows(rejected)

    # — Отчёт —
    removed_total = len(rejected)
    print(f"\n{'─'*62}")
    print(f"  {label}")
    print(f"{'─'*62}")
    print(f"  Входных записей:        {total_in:>8,}")
    print(f"  Закрыто RU-скобок:      {stats.get('ru_paren_fixed',0):>8,}")
    print(f"  Удалено [sq] из EN:     {sum(v for k,v in stats.items() if 'sq_brackets' in k or 'word_brackets' in k):>8,}")
    print(f"  Закрыто EN-скобок:      {sum(v for k,v in stats.items() if 'closed_paren' in k):>8,}")
    print(f"  Исправлено опечаток EN: {sum(v for k,v in stats.items() if 'ocr_fix' in k):>8,}")
    print(f"  Удалено (пустой EN):    {stats.get('empty_after_fix',0):>8,}")
    print(f"  Удалено (adj→compound): {stats.get('bad_tilde_artifact',0):>8,}")
    print(f"  Консолидировано дублей: {dedup_cnt:>8,}")
    print(f"  Итого строк на выходе:  {len(clean_merged):>8,}")

    if rejected:
        print(f"\n  Примеры удалённых (первые 8):")
        for r in rejected[:8]:
            reason = r.get('_reject_reason', '')
            print(f"    [{reason[:50]}]")
            print(f"      EN: {r.get(en_col,'')[:55]}")
            print(f"      RU: {r.get(ru_col,'')[:55]}")

    # Примеры консолидированных синонимов
    multi = [r for r in clean_merged if '; ' in r.get(en_col, '')]
    if multi:
        print(f"\n  Примеры консолидированных синонимов (первые 8):")
        for r in multi[:8]:
            print(f"    RU: {r.get(ru_col,'')[:40]}")
            print(f"    EN: {r.get(en_col,'')[:70]}")

    return {'total_in': total_in, 'clean': len(clean_merged),
            'removed': removed_total, 'dedup': dedup_cnt,
            'rows': clean_merged, 'fields': fields}


# ─────────────────────────────────────────────────────────────────────────────
# 5. XLSX-ГЕНЕРАЦИЯ
# ─────────────────────────────────────────────────────────────────────────────

def save_xlsx(rows, col_headers, xlsx_path, sheet_name='Sheet', hdr_color='1F4E79'):
    if not HAS_XLSX or not rows:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    thin = Side(style='thin', color='CCCCCC')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font  = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        cell.fill  = PatternFill(fill_type='solid', fgColor=hdr_color)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bdr
    ws.row_dimensions[1].height = 20

    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(col_headers, 1):
            val = row.get(col, '')
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical='top', wrap_text=False)
            cell.border = bdr
            if ri % 2 == 0:
                cell.fill = PatternFill(fill_type='solid', fgColor='F2F7FF')

    col_widths = {c: len(c) + 2 for c in col_headers}
    for row in rows[:400]:
        for col in col_headers:
            col_widths[col] = min(70, max(col_widths[col], len(str(row.get(col,''))) + 2))
    for ci, col in enumerate(col_headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[col]

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    wb.save(xlsx_path)
    print(f"  ✓ {os.path.basename(xlsx_path)} → {len(rows):,} строк")


# ─────────────────────────────────────────────────────────────────────────────
# 6. ОБРАБОТКА ВСЕХ ФАЙЛОВ
# ─────────────────────────────────────────────────────────────────────────────

print("═" * 65)
print("FIX ISSUES — финальная правка глоссария")
print("═" * 65)

results = {}

results['ref'] = process_file(
    in_path  = FINAL + r'\reference_glossary_FINAL.tsv',
    out_path = FINAL + r'\reference_glossary_FINAL.tsv',
    label    = 'reference_glossary_FINAL',
)

results['app'] = process_file(
    in_path  = FINAL + r'\approved_glossary_FINAL.tsv',
    out_path = FINAL + r'\approved_glossary_FINAL.tsv',
    label    = 'approved_glossary_FINAL',
)

results['lkp'] = process_file(
    in_path  = FINAL + r'\akzhigitov_lookup.tsv',
    out_path = FINAL + r'\akzhigitov_lookup.tsv',
    label    = 'akzhigitov_lookup',
)


# ─────────────────────────────────────────────────────────────────────────────
# 7. ПЕРЕСОЗДАЁМ XLSX
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "═" * 65)
print("ГЕНЕРАЦИЯ XLSX")
print("═" * 65)

# reference — большой файл
if HAS_XLSX and results.get('ref'):
    r = results['ref']
    ref_cols = [f for f in r['fields'] if f != '_reject_reason']
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reference'
    thin = Side(style='thin', color='CCCCCC')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(ref_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font  = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        cell.fill  = PatternFill(fill_type='solid', fgColor='145A32')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 20
    for ri, row in enumerate(r['rows'], 2):
        for ci, col in enumerate(ref_cols, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col,''))
            cell.alignment = Alignment(vertical='top')
            if ri % 2 == 0:
                cell.fill = PatternFill(fill_type='solid', fgColor='EAFAF1')
    col_w = {'Russian':44,'English':54,'Category':16,'Confidence':20,'Sources':22,'Issues':18}
    for ci, col in enumerate(ref_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_w.get(col, 20)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    wb.save(FINAL + r'\reference_glossary_FINAL.xlsx')
    print(f"  ✓ reference_glossary_FINAL.xlsx → {r['clean']:,} строк")

if results.get('app'):
    save_xlsx(
        rows        = results['app']['rows'],
        col_headers = [f for f in results['app']['fields'] if f != '_reject_reason'],
        xlsx_path   = FINAL + r'\approved_glossary_FINAL.xlsx',
        sheet_name  = 'Approved',
        hdr_color   = '1A5276',
    )

if results.get('lkp'):
    save_xlsx(
        rows        = results['lkp']['rows'],
        col_headers = [f for f in results['lkp']['fields'] if f != '_reject_reason'],
        xlsx_path   = FINAL + r'\akzhigitov_lookup.xlsx',
        sheet_name  = 'Lookup',
        hdr_color   = '784212',
    )


# ─────────────────────────────────────────────────────────────────────────────
# 8. ИТОГОВАЯ СВОДКА
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "═" * 65)
print("ИТОГ FIX ISSUES")
print("═" * 65)

for key, label in [('ref','reference'), ('app','approved'), ('lkp','lookup')]:
    r = results.get(key)
    if r:
        print(f"  {label:<20} {r['total_in']:>7,} → {r['clean']:>7,}"
              f"  (−{r['removed']:,} удалено, −{r['dedup']:,} дубль)")

total = sum(results[k]['clean'] for k in results if results.get(k))
print(f"\n  Итого чистых записей: {total:,}")
print(f"  Файлы: {FINAL}")
