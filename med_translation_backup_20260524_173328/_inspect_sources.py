import sys, csv, json, os
sys.stdout.reconfigure(encoding="utf-8")

BASE = r"C:\Users\Shox\med_translation\словари"
BALDWIN = BASE + r"\baldwin_assets_STRICT_v2_bundle"

# ─── Baldwin STRICT ───────────────────────────────────────────────────
with open(BALDWIN + r"\baldwin_STRICT_safe_glossary_ru_en.tsv", encoding="utf-8-sig") as f:
    bstrict = list(csv.DictReader(f, delimiter="\t"))

print(f"[Baldwin STRICT]  {len(bstrict):,} строк")
print(f"  Колонки : {list(bstrict[0].keys())}")
print(f"  Статусы : {set(r['status'] for r in bstrict)}")
print(f"  Quality : {set(r['quality'] for r in bstrict)}")
for i in [0, 50, 200, 500]:
    r = bstrict[i]
    print(f"  [{i:4}]  {r['source_term']:40s} -> {r['target_term']}")
print()

# ─── Baldwin REFERENCE ────────────────────────────────────────────────
with open(BALDWIN + r"\baldwin_REFERENCE_raw_pairs_ru_en_utf8sig.csv", encoding="utf-8-sig") as f:
    bref = list(csv.DictReader(f))

print(f"[Baldwin REFERENCE]  {len(bref):,} строк")
print(f"  Колонки : {list(bref[0].keys())}")
print(f"  Quality : {set(r['quality'] for r in bref)}")
for i in [0, 50, 200, 500]:
    r = bref[i]
    print(f"  [{i:4}]  {r['source_term']:40s} -> {r['target_term']}")
print()

# ─── Output glossary.tsv (мой extraction) ────────────────────────────
with open(BASE + r"\output\glossary.tsv", encoding="utf-8-sig") as f:
    og = list(csv.DictReader(f, delimiter="\t"))

print(f"[Output glossary.tsv]  {len(og):,} строк")
print(f"  Колонки : {list(og[0].keys())}")
print(f"  Пример  : EN={og[0]['EN']}  RU={og[0]['RU']}  Cat={og[0]['Category']}")
print(f"  Пример  : EN={og[10]['EN']}  RU={og[10]['RU']}")
print()

# ─── Output tm.tsv (мой TM) ──────────────────────────────────────────
with open(BASE + r"\output\tm.tsv", encoding="utf-8-sig") as f:
    tm = list(csv.DictReader(f, delimiter="\t"))

print(f"[Output TM.tsv]  {len(tm):,} строк")
print(f"  Колонки : {list(tm[0].keys())}")
for i in [0, 1, 2]:
    print(f"  [{i}] EN: {tm[i]['EN'][:70]}")
    print(f"       RU: {tm[i]['RU'][:70]}")
print()

# ─── XLSX специализированные ──────────────────────────────────────────
import openpyxl
xlsx_files = [f for f in os.listdir(BASE) if f.endswith(".xlsx")]
print(f"[XLSX файлы]  {len(xlsx_files)} файлов")
for xlsx in sorted(xlsx_files):
    try:
        wb = openpyxl.load_workbook(BASE + "\\" + xlsx, read_only=True, data_only=True)
        for sh in wb.sheetnames:
            ws = wb[sh]
            rows = list(ws.iter_rows(values_only=True, max_row=3))
            print(f"  {xlsx} / [{sh}]  ({ws.max_row} строк, {ws.max_column} колонок)")
            if rows:
                print(f"    Строка 1: {[str(c)[:25] for c in rows[0] if c]}")
            if len(rows) > 1:
                print(f"    Строка 2: {[str(c)[:25] for c in rows[1] if c]}")
        wb.close()
    except Exception as e:
        print(f"  {xlsx}: ERROR {e}")
print()

# ─── PDF разведка ─────────────────────────────────────────────────────
import pdfplumber
pdfs = {
    "2_5.pdf": 5,
    "essential-18000-english-russian-medical-words-dictionary.pdf": 3,
    "covid.pdf": 2,
    "vaccine.pdf": 2,
}
print("[PDF файлы]")
for pdf_name, n_pages in pdfs.items():
    path = BASE + "\\" + pdf_name
    try:
        with pdfplumber.open(path) as pdf:
            total = len(pdf.pages)
            print(f"  {pdf_name}  ({total} стр.)")
            for i in range(min(n_pages, total)):
                text = pdf.pages[i].extract_text() or ""
                lines = [l.strip() for l in text.split("\n") if l.strip()][:5]
                for l in lines:
                    print(f"    [{i+1}] {l[:100]}")
    except Exception as e:
        print(f"  {pdf_name}: ERROR {e}")
