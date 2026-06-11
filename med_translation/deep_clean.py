"""
deep_clean.py — Полная ручная очистка медицинского глоссария.

Стратегия:
  ИСПРАВЛЯЕМ:  EN-префикс в RU-колонке ("dental) зубная..." → "зубная...")
  УДАЛЯЕМ:     тильды, кириллица в EN, OCR-мусор, see-ссылки, артефакты
  ДЕДУПЛИЦИРУЕМ: near-duplicate EN для одного RU (sunstroke / sun stroke)
  ОТЧЁТ:       полный лог отклонённых + статистика причин
"""
import sys, csv, re, os
from collections import defaultdict
sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False
    print("⚠  openpyxl не установлен — XLSX не создаются")

FINAL  = r"C:\Users\Shox\med_translation\final_output"
PARSED = r"C:\Users\Shox\med_translation\словари\parsed"

# ─────────────────────────────────────────────────────────────────────────────
# 1. СИМВОЛЬНЫЕ МНОЖЕСТВА И БЕЛЫЕ СПИСКИ
# ─────────────────────────────────────────────────────────────────────────────

RU_VOWELS     = set('аеёиоуыэюяАЕЁИОУЫЭЮЯ')
# ъ и ь намеренно ИСКЛЮЧЕНЫ: они не образуют консонантных кластеров в начале
RU_CONSONANTS = set('бвгджзйклмнпрстфхцчшщБВГДЖЗЙКЛМНПРСТФХЦЧШЩ')

# Валидные 2-буквенные начальные кластеры согласных русского языка
# (включены все реальные русские слова из медицинской лексики)
RU_VALID_2 = {
    # б-
    'бд','бл','бр','бс','бщ',
    # в-  (вв=введение, вб=вбирать, вж=вживление)
    'вв','вб','вд','вж','вз','вк','вл','вм','вн','вп','вр','вс','вт','вх','вч','вш','вщ',
    # г-  (гв=гвоздь/гвардия)
    'гд','гл','гм','гн','гр','гв',
    # д-
    'дж','дл','дм','дн','дв','др',
    # ж-  (жг=жгут/жгучий, жв=жвачка/жвалы)
    'жд','жж','жг','жв','жн','жл','жр',
    # з-
    'зб','зв','зг','зд','зж','зл','зм','зн','зр','зс',
    # к-  (кз=кзади, кп=кпереди — анатомические направления)
    'кл','кн','кр','кв','кс','кт','кщ','кз','кп',
    # л-  (лж=лживый)
    'лж',
    # м-  (мг=мгновение/мгла, мш=мшистый)
    'мл','мн','мр','мг','мш',
    # н-
    'нж','нк','нр',
    # п-  (пт=птица/птоз/птомаин, пч=пчела, пш=пшеница)
    'пл','пн','пр','пс','пф','пч','пт','пш',
    # р-  (рв=рвота, рт=ртуть, рж=ржавчина)
    'рж','рв','рт',
    # с-  (сс=ссадина/ссора/ссылка)
    'сб','св','сг','сд','сж','сз','сл','см','сн','сп','ср','сс','ст','сф','сх','сц',
    'сч','ск','сш',
    # т-  (тщ=тщательный)
    'тв','тк','тл','тм','тр','тс','тч','тщ',
    # ф-  (фт=фтор/фтизиатрия)
    'фл','фр','фт',
    # х-
    'хл','хм','хн','хр','хв','хт',
    # ц-
    'цв','цк',
    # ч-
    'чл','чт','чр',
    # ш-
    'шл','шм','шн','шп','шр','шт','шк','шв','шц',
    # щ-
    'щт','щр',
}

# Валидные 3-буквенные начальные кластеры
RU_VALID_3 = {
    # вс* — вст=встреча, вск=вскрыть, всп=всплеск/вспышка
    'вст','вск','всп',
    # вз* — взр=взрыв, взд=вздохнуть, взм=взмах, взл=взлёт, взг=взгляд
    'взр','взд','взм','взл','взн','взб','взв','взг',
    # стр, скр, спр и т.д.
    'стр','скр','спр','спл','сдр','сбр','стл','скл','сбл','ств','сгр','сфр','сшт',
    'шпр','шкр','хвр',
}


# ─────────────────────────────────────────────────────────────────────────────
# 2. ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ─────────────────────────────────────────────────────────────────────────────

def cyrillic_ratio(s):
    letters = re.findall(r'[A-Za-zА-ЯЁа-яё]', s)
    if not letters: return 0.0
    return sum(1 for c in letters if re.match(r'[А-ЯЁа-яё]', c)) / len(letters)

def latin_ratio(s):
    letters = re.findall(r'[A-Za-zА-ЯЁа-яё]', s)
    if not letters: return 0.0
    return sum(1 for c in letters if re.match(r'[A-Za-z]', c)) / len(letters)

def has_cyrillic(s): return bool(re.search(r'[А-ЯЁа-яё]', s))
def has_latin(s):    return bool(re.search(r'[A-Za-z]', s))

def consonant_cluster_start(word):
    """Начальный кластер согласных (только из RU_CONSONANTS, без ъ/ь)."""
    cluster = ''
    for c in word.lower():
        if c in {x.lower() for x in RU_CONSONANTS}:
            cluster += c
        else:
            break
    return cluster

def is_valid_word_start(word):
    """True если начало слова соответствует русской фонотактике."""
    cluster = consonant_cluster_start(word)
    n = len(cluster)
    if n <= 1: return True
    if n == 2: return cluster in RU_VALID_2
    if n == 3: return cluster[:2] in RU_VALID_2 or cluster[:3] in RU_VALID_3
    # 4-согласных: вскр=вскрыть, встр=встреча, вздр=вздрогнуть, взгл=взглянуть
    # Проверяем первые 3 символа кластера
    if n == 4: return cluster[:3] in RU_VALID_3
    # 5+ согласных подряд → всегда мусор
    return False

def has_no_vowels_ru(word):
    """Кириллическое слово 5+ символов без гласных."""
    ru_chars = [c for c in word
                if c.lower() in {v.lower() for v in RU_VOWELS | RU_CONSONANTS}]
    return len(ru_chars) >= 5 and not any(c in RU_VOWELS for c in word)

def has_digit_in_cyr_word(s):
    """Цифра внутри кириллического слова: алл01снная."""
    return bool(re.search(r'[А-ЯЁа-яё]\d[А-ЯЁа-яё]', s))

def text_is_ocr_garbage(s):
    """Возвращает (is_garbage: bool, reason: str | None)."""
    if has_digit_in_cyr_word(s):
        return True, 'digit_in_cyr_word'

    cyr_words = re.findall(r'[А-ЯЁа-яё]{5,}', s)
    if not cyr_words:
        return False, None

    bad = []
    for w in cyr_words:
        if has_no_vowels_ru(w):
            bad.append((w, 'no_vowels'))
        elif not is_valid_word_start(w):
            bad.append((w, 'bad_cluster'))

    if bad and len(bad) / len(cyr_words) >= 0.40:
        return True, 'ocr:' + ','.join(b[0] for b in bad[:2])

    return False, None


# ─────────────────────────────────────────────────────────────────────────────
# 3. ИСПРАВЛЕНИЕ И ПРОВЕРКА ПОЛЕЙ
# ─────────────────────────────────────────────────────────────────────────────

# EN-контекст-префикс в RU: "dental) зубная..." → "зубная..."
_RU_EN_PREFIX = re.compile(
    r'^[a-zA-Z][a-zA-Z\s,/\-\.]{1,30}\)\s+(?=[А-ЯЁа-яё])'
)
# "(Dental) зубная..." → "зубная..."
_RU_PAREN_EN  = re.compile(r'^\([a-zA-Z][^\)]{1,30}\)\s+')


def fix_ru(ru):
    """Исправляем RU: убираем English-префиксы.
    Возвращает (fixed_ru, fix_description | None).
    """
    changes = []

    m = _RU_EN_PREFIX.match(ru)
    if m:
        prefix = m.group().strip()
        ru = ru[m.end():]
        changes.append(f'stripped_prefix:{prefix!r}')

    m2 = _RU_PAREN_EN.match(ru)
    if m2 and re.search(r'[A-Za-z]', m2.group()):
        paren = m2.group().strip()
        ru = ru[m2.end():]
        changes.append(f'stripped_paren:{paren!r}')

    ru = re.sub(r'[\s\-–—,;\.]+$', '', ru).strip()
    ru = re.sub(r'^[\s,;\.]+', '', ru).strip()

    return ru, ('; '.join(changes) if changes else None)


def check_en(en):
    """(is_bad: bool, reason: str) для EN-поля."""
    en = en.strip()
    if not en or len(en) < 2:
        return True, 'empty_en'
    if '~' in en:
        return True, 'tilde_in_en'
    if cyrillic_ratio(en) > 0.08:
        return True, f'cyr_in_en={cyrillic_ratio(en):.2f}'
    if re.match(r'^[\d\W\s]+$', en):
        return True, 'no_letters_en'
    is_g, r = text_is_ocr_garbage(en)
    if is_g:
        return True, f'en_ocr:{r}'
    return False, None


def check_ru(ru):
    """(is_bad: bool, reason: str) для RU-поля (после исправления)."""
    ru = ru.strip()
    if not ru or len(ru) < 2:
        return True, 'empty_ru'
    if '~' in ru:
        return True, 'tilde_in_ru'
    if re.match(r'^см\.', ru, re.IGNORECASE):
        return True, 'see_ref'
    if re.match(r'^see\b', ru, re.IGNORECASE):
        return True, 'see_ref_en'
    # Высокая латиница: содержит настоящие EN-слова (не просто аббревиатуры)
    if latin_ratio(ru) > 0.40:
        return True, f'high_latin={latin_ratio(ru):.2f}'
    is_g, r = text_is_ocr_garbage(ru)
    if is_g:
        return True, f'ru_ocr:{r}'
    return False, None


# ─────────────────────────────────────────────────────────────────────────────
# 4. ДЕДУПЛИКАЦИЯ near-duplicate EN для одного RU
# ─────────────────────────────────────────────────────────────────────────────

def normalize_en_for_dedup(en):
    """Нормализуем EN: нижний регистр, убираем пробелы и знаки."""
    return re.sub(r'[^a-z0-9]', '', en.lower())


def dedup_synonyms(rows, ru_col='Russian', en_col='English'):
    """
    Для каждого RU-термина удаляем near-duplicate EN.
    'sunstroke' / 'sun stroke' → оставляем первое встреченное.
    Возвращает (clean_rows, dedup_count).
    """
    ru_groups = defaultdict(list)
    for r in rows:
        ru_groups[r.get(ru_col, '').lower().strip()].append(r)

    result = []
    dedup_count = 0
    for _, group in ru_groups.items():
        seen = {}
        for r in group:
            en_norm = normalize_en_for_dedup(r.get(en_col, ''))
            if not en_norm:
                continue
            if en_norm in seen:
                dedup_count += 1
            else:
                seen[en_norm] = True
                result.append(r)

    # Сохраняем исходный порядок (по Russian)
    result.sort(key=lambda r: r.get(ru_col, '').lower())
    return result, dedup_count


# ─────────────────────────────────────────────────────────────────────────────
# 5. ОСНОВНАЯ ФУНКЦИЯ ОБРАБОТКИ ФАЙЛА
# ─────────────────────────────────────────────────────────────────────────────

def process_file(in_path, out_path, reject_path,
                 ru_col='Russian', en_col='English', label=''):
    if not os.path.exists(in_path):
        print(f"  ⚠  Файл не найден: {in_path}")
        return None

    with open(in_path, encoding='utf-8-sig') as f:
        rows = list(csv.DictReader(f, delimiter='\t'))

    if not rows:
        print(f"  ⚠  Файл пустой: {in_path}")
        return None

    fields = list(rows[0].keys())
    total_in = len(rows)

    stats     = defaultdict(int)
    clean     = []
    rejected  = []
    fixed     = 0

    for r in rows:
        ru = r.get(ru_col, '').strip()
        en = r.get(en_col, '').strip()

        if not ru or not en:
            stats['empty_field'] += 1
            r['_reject_reason'] = 'empty_field'
            rejected.append(r)
            continue

        # — Исправляем RU (EN-префикс) —
        ru_new, fix_desc = fix_ru(ru)
        if fix_desc:
            r[ru_col] = ru_new
            ru = ru_new
            fixed += 1
            stats['fixed'] += 1

        # — Проверяем EN —
        en_bad, en_reason = check_en(en)
        if en_bad:
            stats[f'EN:{en_reason}'] += 1
            r['_reject_reason'] = f'EN: {en_reason}'
            rejected.append(r)
            continue

        # — Проверяем RU (после исправления) —
        ru_bad, ru_reason = check_ru(ru)
        if ru_bad:
            stats[f'RU:{ru_reason}'] += 1
            r['_reject_reason'] = f'RU: {ru_reason}'
            rejected.append(r)
            continue

        # — Минимальная осмысленная длина (аббревиатуры типа Гц/МУ допустимы) —
        if len(re.sub(r'[^\w]', '', ru)) < 2 or len(re.sub(r'[^\w]', '', en)) < 2:
            stats['too_short'] += 1
            r['_reject_reason'] = 'too_short'
            rejected.append(r)
            continue

        clean.append(r)
        stats['ok'] += 1

    # — Дедупликация near-duplicate EN —
    clean_dd, dedup_cnt = dedup_synonyms(clean, ru_col, en_col)
    stats['dedup_en_synonyms'] = dedup_cnt

    # — Сохраняем чистый файл —
    with open(out_path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter='\t', extrasaction='ignore')
        w.writeheader()
        w.writerows(clean_dd)

    # — Сохраняем отклонённые —
    reject_fields = fields + ['_reject_reason']
    with open(reject_path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.DictWriter(f, fieldnames=reject_fields, delimiter='\t', extrasaction='ignore')
        w.writeheader()
        w.writerows(rejected)

    # — Выводим статистику —
    removed = total_in - len(clean_dd)
    print(f"\n{'─'*60}")
    print(f"  {label}")
    print(f"{'─'*60}")
    print(f"  Входных записей:       {total_in:>8,}")
    print(f"  Исправлено (EN-prefix):{fixed:>8,}")
    print(f"  Удалено всего:         {len(rejected):>8,}  ({len(rejected)/total_in*100:.1f}%)")
    print(f"  Дедупликация EN:       {dedup_cnt:>8,}")
    print(f"  Итого чистых:          {len(clean_dd):>8,}")

    print(f"\n  Причины удаления:")
    for k, v in sorted(stats.items(), key=lambda x: -x[1]):
        if k not in ('ok', 'fixed', 'dedup_en_synonyms') and v > 0:
            print(f"    {k:<35} {v:>6,}")

    # — Примеры удалённых —
    if rejected:
        print(f"\n  Примеры удалённых (первые 8):")
        for r in rejected[:8]:
            reason = r.get('_reject_reason', '')
            en = r.get(en_col, '')[:45]
            ru = r.get(ru_col, '')[:55]
            print(f"    [{reason}]")
            print(f"      EN: {en}")
            print(f"      RU: {ru}")

    return {'total_in': total_in, 'fixed': fixed,
            'removed': len(rejected), 'dedup': dedup_cnt,
            'clean': len(clean_dd), 'rows': clean_dd, 'fields': fields}


# ─────────────────────────────────────────────────────────────────────────────
# 6. XLSX-ГЕНЕРАЦИЯ
# ─────────────────────────────────────────────────────────────────────────────

def save_xlsx(rows, col_headers, xlsx_path, sheet_name='Sheet',
              hdr_color='1F4E79', even_color='F2F7FF'):
    if not HAS_XLSX or not rows:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    thin = Side(style='thin', color='CCCCCC')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Заголовок
    for ci, col in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        cell.fill      = PatternFill(fill_type='solid', fgColor=hdr_color)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = bdr
    ws.row_dimensions[1].height = 20

    # Данные
    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(col_headers, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col, ''))
            cell.alignment = Alignment(vertical='top', wrap_text=False)
            cell.border    = bdr
            if ri % 2 == 0:
                cell.fill = PatternFill(fill_type='solid', fgColor=even_color)

    # Ширина колонок (семплируем первые 300 строк)
    col_widths = {c: len(c) + 2 for c in col_headers}
    for row in rows[:300]:
        for col in col_headers:
            col_widths[col] = min(62, max(col_widths[col], len(str(row.get(col,''))) + 2))

    for ci, col in enumerate(col_headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[col]

    ws.freeze_panes  = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(xlsx_path)
    print(f"  ✓ {os.path.basename(xlsx_path)} → {len(rows):,} строк")


# ─────────────────────────────────────────────────────────────────────────────
# 7. ОБРАБОТКА ВСЕХ ФАЙЛОВ
# ─────────────────────────────────────────────────────────────────────────────

print("═" * 65)
print("DEEP CLEAN — полная очистка медицинского глоссария")
print("═" * 65)

results = {}

# 7a. reference_glossary_FINAL
results['ref'] = process_file(
    in_path     = FINAL + r'\reference_glossary_FINAL.tsv',
    out_path    = FINAL + r'\reference_glossary_FINAL.tsv',
    reject_path = FINAL + r'\rejected_deepclean_reference.tsv',
    ru_col='Russian', en_col='English',
    label='reference_glossary_FINAL'
)

# 7b. approved_glossary_FINAL
results['app'] = process_file(
    in_path     = FINAL + r'\approved_glossary_FINAL.tsv',
    out_path    = FINAL + r'\approved_glossary_FINAL.tsv',
    reject_path = FINAL + r'\rejected_deepclean_approved.tsv',
    ru_col='Russian', en_col='English',
    label='approved_glossary_FINAL'
)

# 7c. akzhigitov_lookup
results['lkp'] = process_file(
    in_path     = FINAL + r'\akzhigitov_lookup.tsv',
    out_path    = FINAL + r'\akzhigitov_lookup.tsv',
    reject_path = FINAL + r'\rejected_deepclean_lookup.tsv',
    ru_col='Russian', en_col='English',
    label='akzhigitov_lookup'
)


# ─────────────────────────────────────────────────────────────────────────────
# 8. ПЕРЕСОЗДАЁМ XLSX ДЛЯ ВСЕХ ФАЙЛОВ
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "═" * 65)
print("ГЕНЕРАЦИЯ XLSX")
print("═" * 65)

if results.get('ref'):
    ref_cols = ['Russian', 'English', 'Category', 'Confidence', 'Sources', 'Issues']
    # Для большого файла — цикл построчно
    r = results['ref']
    if HAS_XLSX:
        xlsx_path = FINAL + r'\reference_glossary_FINAL.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reference'
        thin = Side(style='thin', color='CCCCCC')
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ci, col in enumerate(ref_cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font  = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
            cell.fill  = PatternFill(fill_type='solid', fgColor='145A32')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 20
        for ri, row in enumerate(r['rows'], 2):
            for ci, col in enumerate(ref_cols, 1):
                cell = ws.cell(row=ri, column=ci, value=row.get(col,''))
                cell.alignment = Alignment(vertical='top')
                if ri % 2 == 0:
                    cell.fill = PatternFill(fill_type='solid', fgColor='EAFAF1')
        col_widths_ref = {'Russian':42,'English':42,'Category':16,
                          'Confidence':20,'Sources':22,'Issues':18}
        for ci, col in enumerate(ref_cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = col_widths_ref.get(col,20)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        wb.save(xlsx_path)
        print(f"  ✓ reference_glossary_FINAL.xlsx → {r['clean']:,} строк")

if results.get('app'):
    save_xlsx(
        rows        = results['app']['rows'],
        col_headers = [f for f in results['app']['fields'] if f != '_reject_reason'],
        xlsx_path   = FINAL + r'\approved_glossary_FINAL.xlsx',
        sheet_name  = 'Approved',
        hdr_color   = '1A5276',
    )

if results.get('lkp'):
    save_xlsx(
        rows        = results['lkp']['rows'],
        col_headers = [f for f in results['lkp']['fields'] if f != '_reject_reason'],
        xlsx_path   = FINAL + r'\akzhigitov_lookup.xlsx',
        sheet_name  = 'Lookup',
        hdr_color   = '784212',
    )

# Parsed source files
with open(PARSED + r'\2_5_pairs.tsv', encoding='utf-8-sig') as f:
    akzh_rows = list(csv.DictReader(f, delimiter='\t'))
akzh_cols = list(akzh_rows[0].keys()) if akzh_rows else ['English','Russian']
save_xlsx(akzh_rows, akzh_cols,
          FINAL + r'\2_5_glossary.xlsx', 'Akzhigitov', '4A235A')

with open(PARSED + r'\covid_pairs.tsv', encoding='utf-8-sig') as f:
    covid_rows = list(csv.DictReader(f, delimiter='\t'))
covid_cols = list(covid_rows[0].keys()) if covid_rows else ['English','Russian']
save_xlsx(covid_rows, covid_cols,
          FINAL + r'\covid_glossary.xlsx', 'COVID', '922B21')


# ─────────────────────────────────────────────────────────────────────────────
# 9. ИТОГОВАЯ СВОДКА
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "═" * 65)
print("ИТОГ DEEP CLEAN")
print("═" * 65)

for key, label in [('ref','reference'), ('app','approved'), ('lkp','lookup')]:
    r = results.get(key)
    if r:
        pct = r['removed']/r['total_in']*100 if r['total_in'] else 0
        print(f"  {label:<20} {r['total_in']:>7,} → {r['clean']:>7,}"
              f"  (−{r['removed']:,} удалено {pct:.1f}%,"
              f" −{r['dedup']:,} дубль, +{r['fixed']:,} исправлено)")

total_clean = sum(results[k]['clean'] for k in results if results.get(k))
print(f"\n  Итого чистых записей: {total_clean:,}")
print(f"\n  Логи отклонённых:")
print(f"    {FINAL}\\rejected_deepclean_reference.tsv")
print(f"    {FINAL}\\rejected_deepclean_approved.tsv")
print(f"    {FINAL}\\rejected_deepclean_lookup.tsv")
print(f"\n  XLSX обновлены в: {FINAL}")
